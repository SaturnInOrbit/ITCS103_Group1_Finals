import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os
from openpyxl.styles import Font

EXCEL_FILE = "inventory.xlsx"

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Name", "Price", "Stock", "Image"])
        wb.save(EXCEL_FILE)

ensure_excel_file()

# Setup
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")  # We'll manually adjust purple later

# Constants
ADMIN_USER = "admin"
ADMIN_PASS = "password"
FILE_NAME = "users.xlsx"
# Global storage
inventory = []
sales = []
purchase_history = []
user_accounts = {}  # Dictionary to store user credentials and data

nav_buttons = []



# Product structure
class Product:
    def __init__(self, name, price, stock, image_path):
        self.name = name
        self.price = price
        self.stock = stock
        self.image_path = image_path

class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.cart = []
        self.purchase_history = []

def validate_login(username, password):
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                return True
        return False

# Main App
class ECommerceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("E-Commerce Inventory System")
        self.geometry("1000x700")
        self.resizable(False, False)
        self.configure(bg="#1a1a1a")

        self.error_label = None
        self.selected_product = None
        self.current_user = None

        self.home_screen()

    def clear_widgets(self):
        for widget in self.winfo_children():
            widget.destroy()

    def home_screen(self):
        self.clear_widgets()

        # Main Frame
        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        # Title Label with updated font and color
        title = ctk.CTkLabel(frame, text="Welcome to the E-Commerce System", font=("Arial", 34, "bold"), text_color="#b84dff")
        title.pack(pady=30)

        # Admin Login Button with modern styling
        admin_btn = ctk.CTkButton(frame, text="👨‍💼 Admin Login", command=self.admin_login_screen, 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 16, "bold"))
        admin_btn.pack(pady=20, ipadx=30, ipady=10, fill="x", padx=40)

        # Shop Now Button with modern styling and added icon
        user_btn = ctk.CTkButton(frame, text="🛒 Shop Now", command=self.user_login_screen, 
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"))
        user_btn.pack(pady=20, ipadx=30, ipady=10, fill="x", padx=40)

        quit_btn = ctk.CTkButton(frame, text="Exit", command=self.destroy, 
                                fg_color="#FF0000", hover_color="#800000", font=("Arial", 16, "bold"))
        quit_btn.pack(pady=20, ipadx=30, ipady=10, fill="x", padx=100)

    def admin_login_screen(self):
        self.clear_widgets()

        # Main Frame
        frame = ctk.CTkFrame(self, fg_color="#2a2a2a")
        frame.pack(expand=True, padx=40, pady=40)

        # Title with updated font size and color
        title = ctk.CTkLabel(frame, text="👨‍💼 Admin Login", font=("Arial", 26, "bold"), text_color="#b84dff")
        title.pack(pady=20,padx = 20)

        # Username Label and Entry
        user_label = ctk.CTkLabel(frame, text="Username:", font=("Arial", 14))
        user_label.pack(pady=(10, 0))
        self.admin_username = ctk.CTkEntry(frame, placeholder_text="Enter your username", font=("Arial", 14))
        self.admin_username.pack(pady=(0, 10),padx =10, fill="x")

        # Password Label and Entry
        pass_label = ctk.CTkLabel(frame, text="Password:", font=("Arial", 14))
        pass_label.pack(pady=(10, 0))
        self.admin_password = ctk.CTkEntry(frame, show="*", placeholder_text="Enter your password", font=("Arial", 14))
        self.admin_password.pack(pady=(0, 20),padx =10, fill="x")

        # Login Button with modern design
        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_admin, 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 16, "bold"), border_width=2, corner_radius=10)
        login_btn.pack(pady=(0, 10),padx= 10,  fill="x")

        # Back Button with updated design
        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen, 
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"), border_width=2, corner_radius=10)
        back_btn.pack(pady=10,padx =10, fill="x")


    def verify_admin(self):
        username = self.admin_username.get()
        password = self.admin_password.get()

        if username == ADMIN_USER and password == ADMIN_PASS:
            self.admin_dashboard()
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
        else:
            messagebox.showerror("Login Failed", "Invalid admin credentials")
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
    
    def admin_logout(self):
        msglogout = messagebox.askokcancel("Log Out", "Are you sure you're logging out?")
        if msglogout == False:
            pass
        else: 
            self.home_screen()
    
# --- ADMIN DASHBOARD ---
    def admin_dashboard(self):
        self.clear_widgets()

        # Main layout: sidebar and content area
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill="both", expand=True)

        # Sidebar
        sidebar = ctk.CTkFrame(main_frame, width=200, fg_color="#1a1a1a")
        sidebar.pack(side="left", fill="y")

        title = ctk.CTkLabel(sidebar, text="👨‍💼 Admin", font=("Arial", 24, "bold"), text_color="#ff9933")
        title.pack(pady=(20, 30))

        # Sidebar buttons
        add_product_btn = ctk.CTkButton(sidebar, text="➕ Add Product", command=self.add_product_popup)
        add_product_btn.pack(pady=10, fill="x", padx=10)

        download_sales_btn = ctk.CTkButton(sidebar, text="📥 Download Sales", command=self.download_sales_report)
        download_sales_btn.pack(pady=10, fill="x", padx=10)

        logout_btn = ctk.CTkButton(sidebar, text="🚪 Logout", command=self.admin_logout)
        logout_btn.pack(pady=20, fill="x", padx=10)

        # Content Area
        self.admin_frame = ctk.CTkFrame(main_frame, fg_color="#262626")
        self.admin_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        self.refresh_inventory()

    def refresh_inventory(self):
        from PIL import Image, ImageTk
        import os

        # Clear previous widgets
        for widget in self.admin_frame.winfo_children():
            widget.destroy()

        # Clear current inventory list
        inventory.clear()

        # Load from Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, stock, image_path = row
            if os.path.exists(image_path):  # Ensure image still exists
                product = Product(name, float(price), int(stock), image_path)
                inventory.append(product)

        if not inventory:
            label = ctk.CTkLabel(self.admin_frame, text="No Products Yet.", text_color="gray")
            label.pack()
            return

        # Display products
        for idx, product in enumerate(inventory):
            frame = ctk.CTkFrame(self.admin_frame, fg_color="#333333", corner_radius=10)
            frame.grid(row=idx//2, column=idx%2, padx=10, pady=10, sticky="nsew")

            try:
                img = Image.open(product.image_path)
                img = img.resize((100, 100))
                img = ImageTk.PhotoImage(img)
                img_label = ctk.CTkLabel(frame, image=img, text="")
                img_label.image = img  # keep reference
                img_label.pack()
            except Exception as e:
                error_label = ctk.CTkLabel(frame, text="Image error", text_color="red")
                error_label.pack()

            info = f"{product.name}\n₱{product.price:.2f}\nStock: {product.stock}"
            label = ctk.CTkLabel(frame, text=info)
            label.pack()

            edit_btn = ctk.CTkButton(frame, text="Edit Stock", command=lambda p=product: self.edit_stock_popup(p))
            edit_btn.pack(pady=2)   

            delete_btn = ctk.CTkButton(frame, text="Delete", fg_color="red", command=lambda p=product: self.delete_product(p))
            delete_btn.pack(pady=2)


    def add_product_popup(self):
        # Create the popup window
        popup = tk.Toplevel(self)
        popup.title("Add Product")
        popup.geometry("350x700")
        popup.configure(bg="#2a2a2a")  # Darker background for better contrast

        # Product Name Label and Entry
        name_label = ctk.CTkLabel(popup, text="Product Name:", font=("Arial", 14), text_color="#b84dff")
        name_label.pack(pady=(20, 5))
        name_entry = ctk.CTkEntry(popup, placeholder_text="Enter product name", font=("Arial", 14))
        name_entry.pack(pady=(0, 10),padx=10, fill="x")

        # Price Label and Entry
        price_label = ctk.CTkLabel(popup, text="Price:", font=("Arial", 14), text_color="#b84dff")
        price_label.pack(pady=(10, 5))
        price_entry = ctk.CTkEntry(popup, placeholder_text="Enter product price", font=("Arial", 14))
        price_entry.pack(pady=(0, 10),padx=10, fill="x")

        # Stock Label and Entry
        stock_label = ctk.CTkLabel(popup, text="Stock:", font=("Arial", 14), text_color="#b84dff")
        stock_label.pack(pady=(10, 5))
        stock_entry = ctk.CTkEntry(popup, placeholder_text="Enter stock quantity", font=("Arial", 14))
        stock_entry.pack(pady=(0, 10),padx =10, fill="x") 

        # Image Button
        image_btn = ctk.CTkButton(popup, text="➕ Select Image", command=lambda: self.select_image(popup), 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 14))
        image_btn.pack(pady=(10, 20),padx = 10, fill="x")

        # Add Product Button
        add_btn = ctk.CTkButton(popup, text="Add Product", command=lambda: self.save_product(name_entry.get(), 
                                                                                            price_entry.get(), 
                                                                                            stock_entry.get(), 
                                                                                            self.selected_image, 
                                                                                            popup),
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"))
        add_btn.pack(pady=(10, 20),padx =10, ipadx=30, ipady=10, fill="x")

        # Back Button (optional)
        back_btn = ctk.CTkButton(popup, text="Back", command=popup.destroy, 
                                fg_color="#ff3333", hover_color="#cc2900", font=("Arial", 14, "bold"))
        back_btn.pack(pady=(10, 20), ipadx=10, ipady=10, fill="x")


    def select_image(self, popup):
        path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg")])
        if path:
            self.selected_image = path
            messagebox.showinfo("Image Selected", "Image successfully selected.")

    def save_product(self, name, price, stock, image, popup):
        if not (name and price and stock and image):
            messagebox.showerror("Error", "All fields are required.")
            return
        try:
            price = float(price)
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Price must be a number. Stock must be an integer.")
            return

        new_product = Product(name, price, stock, image)
        inventory.append(new_product)

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]
        ws.append([name, price, stock, image])
        wb.save(EXCEL_FILE)

        popup.destroy()
        self.refresh_inventory()

    def edit_stock_popup(self, product):
        popup = tk.Toplevel(self)
        popup.title("Edit Stock")
        popup.geometry("250x400")
        popup.configure(bg="#1a1a1a")

        # Display the current product name and stock
        label = ctk.CTkLabel(popup, text=f"Editing stock for {product.name}")
        label.pack(pady=10)

        # Entry field for updating stock
        stock_entry = ctk.CTkEntry(popup)
        stock_entry.insert(0, str(product.stock))  # Pre-fill with current stock
        stock_entry.pack(pady=10)

        price_label = ctk.CTkLabel(popup, text="Price (₱):")
        price_label.pack()
        price_entry = ctk.CTkEntry(popup)
        price_entry.insert(0, f"{product.price:.2f}")
        price_entry.pack(pady=5)

        # Save button that will call save_stock function to update the stock in the Excel sheet
        save_btn = ctk.CTkButton(popup, text="Save", command=lambda: self.save_stock(product, stock_entry.get(), price_entry.get(), popup))
        save_btn.pack(pady=10)

    def save_stock(self, product, new_stock, new_price, popup):
        # Convert the new stock value to integer
        try:
            new_stock = int(new_stock)
            new_price = float(new_price)
        except ValueError:
            messagebox.showerror("Invalid Input", "Please enter a valid number for stock and price.")
            return

        # Update the stock value in the Excel file
        try:
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Products"]
            
            # Find the row where the product is located
            for row in ws.iter_rows(min_row=2):
                if row[0].value == product.name:  # Match the product name
                    row[1].value = new_price
                    row[2].value = new_stock  # Update the stock in column C (index 2)

                    break
            
            # Save the updated Excel file
            wb.save(EXCEL_FILE)

            # Close the popup after saving
            popup.destroy()

            messagebox.showinfo("Success", "Stock updated successfully.")
            self.refresh_inventory()
        
        except Exception as e:
            messagebox.showerror("Error", f"Failed to update stock: {e}")

    def delete_product(self, product): ##not working
        if messagebox.askyesno("Delete", f"Are you sure you want to delete {product.name}?"):
            # Remove from in-memory list
            if product in inventory:
                inventory.remove(product)

            # Remove from Excel
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Products"]

            for row in list(ws.iter_rows(min_row=2)):
                excel_name = str(row[0].value).strip()
                excel_price = float(row[1].value)
                excel_stock = int(row[2].value)
                excel_image = str(row[3].value).strip() if row[3].value else ""

                if (excel_name.lower() == product.name.lower() and
                    excel_price == product.price and
                    excel_stock == product.stock and
                    excel_image == product.image_path.strip()):
                    
                    ws.delete_rows(row[0].row)
                    break

            wb.save(EXCEL_FILE)

            # Optional: delete image file if it exists
            try:
                if os.path.exists(product.image_path):
                    os.remove(product.image_path)
            except Exception as e:
                print("Warning: Could not delete image file:", e)

            # Refresh product display
            self.refresh_inventory()

    def download_sales_report(self):
        if not sales:
            messagebox.showerror("Error", "No sales to report.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        headers = ["Username", "Product", "Price", "Date"]
        ws.append(headers)

        # Append each sale in the sales list to the Excel file
        for sale in sales:
            ws.append(sale)

        wb.save(filepath)
        messagebox.showinfo("Success", f"Sales report saved to {filepath}.")

    def user_login_screen(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        
        title = ctk.CTkLabel(frame, text="User Login", font=("Arial", 26, "bold"), text_color="#00cc66")
        title.pack(pady=(10, 20))

        # Username fiel
        user_label = ctk.CTkLabel(frame, text="Username:", text_color="white", font=("Arial", 14))
        user_label.pack(anchor="w", padx=10)
        self.user_username = ctk.CTkEntry(frame, width=250, font=("Arial", 14), fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.user_username.pack(pady=(0, 10), padx=10)

        # Password fieldd
        pass_label = ctk.CTkLabel(frame, text="Password:", text_color="white", font=("Arial", 14))
        pass_label.pack(anchor="w", padx=10)
        self.user_password = ctk.CTkEntry(frame, width=250, font=("Arial", 14), show="*", fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.user_password.pack(pady=(0, 20), padx=10)

        # Login Button
        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_user, font=("Arial", 14), fg_color="#00cc66", hover_color="#00994d", width=250, corner_radius=10)
        login_btn.pack(pady=(10, 10), padx=10)

        # Register Button (if user doesn't have an account yet)
        register_btn = ctk.CTkButton(frame, text="Register", command=self.user_register_screen, font=("Arial", 14), fg_color="#b84dff", hover_color="#9933cc", width=250, corner_radius=10)
        register_btn.pack(pady=(5, 10), padx=10)

        # Back Button to tangina gulo nad
        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen, font=("Arial", 14), fg_color="#ff6666", hover_color="#ff4d4d", width=250, corner_radius=10)
        back_btn.pack(pady=(5, 10), padx=10)

    
    def verify_user(self):
        username = self.user_username.get()
        password = self.user_password.get()

        if validate_login(username, password):
            messagebox.showinfo("Login Success!", f"Welcome, {username}!")
            self.current_user = User(username, password)
            self.user_window()
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
        else:
            messagebox.showerror("Login Failed", "Invalid username or password.")
            self.user_username.delete(0, "end")
            self.user_password.delete(0, "end")

    def user_register_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="User Registration", font=("Arial", 26, "bold"), text_color="#00cc66")
        title.pack(pady=20)

        user_label = ctk.CTkLabel(frame, text="New Username:", font=("Arial", 14), text_color="white")
        user_label.pack(anchor="w", padx=10, pady=(10, 0))
        self.reg_username = ctk.CTkEntry(frame, width=250, font=("Arial", 14), fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.reg_username.pack(pady=(0, 10), padx=10)

        pass_label = ctk.CTkLabel(frame, text="New Password:", font=("Arial", 14), text_color="white")
        pass_label.pack(anchor="w", padx=10, pady=(10, 0))
        self.reg_password = ctk.CTkEntry(frame, width=250, font=("Arial", 14), show="*", fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.reg_password.pack(pady=(0, 10), padx=10)

        register_btn = ctk.CTkButton(frame, text="Register", command=self.save_register_user, font=("Arial", 14), fg_color="#00cc66", hover_color="#00994d", width=250, corner_radius=10)
        register_btn.pack(pady=(10, 10), padx=10)

        back_btn = ctk.CTkButton(frame, text="Back", command=self.user_login_screen, font=("Arial", 14), fg_color="#ff6666", hover_color="#ff4d4d", width=250, corner_radius=10)
        back_btn.pack(pady=(5, 10), padx=10)


    def register_user(self):
        if os.path.exists(FILE_NAME):
            wb = openpyxl.load_workbook(FILE_NAME)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "accounts"
            ws.append(["username", "password"])
            for col in range(1, 3):  # Fix range here
                ws.cell(row=1, column=col).font = Font(bold=True)
        return wb, ws

    def save_register_user(self):
        username = self.reg_username.get()
        password = self.reg_password.get()

        if not username or not password:
            messagebox.showerror("error","plesase fill up the entries")

        wb, ws = self.register_user()
        ws.append([username, password])
        wb.save(FILE_NAME)
          
        self.reg_username.delete(0, "end")
        self.reg_password.delete(0, "end")
        messagebox.showinfo('Success!',f'Username: {username} registered successfully!')
        self.user_login_screen()

    def user_window(self):
        self.clear_widgets()

        title_label = ctk.CTkLabel(self, text="🛍️ The One-stop Shop for all your daily needs!", font=ctk.CTkFont(size=22, weight="bold"))
        title_label.pack(pady=(15, 5))

        # ---- Toolbar Frame (Horizontal Buttons) ----
        toolbar = ctk.CTkFrame(self)
        toolbar.pack(fill="x", pady=5)

        # ---- Content Area ----
        self.content_area = ctk.CTkFrame(self)
        self.content_area.pack(fill="both", expand=True, padx=20, pady=20)

        self.content_label = ctk.CTkLabel(self.content_area, text="Welcome to the User Dashboard!", font=ctk.CTkFont(size=18))
        self.content_label.pack(pady=20)
        self.content_label2 = ctk.CTkLabel(self.content_area, text="Press any of the top buttons to begin!", font=ctk.CTkFont(size=18))
        self.content_label2.pack(pady=20)

        def clear_content():
            for widget in self.content_area.winfo_children():
                widget.destroy()

        def set_active(button):
            for btn in nav_buttons:
                
                btn.configure(fg_color="transparent")
            button.configure(fg_color="#11aa11")

        def show_products():
            set_active(product_btn)
            clear_content()

            self.content_label = ctk.CTkLabel(self.content_area, text="📦 Products Page", font=ctk.CTkFont(size=18))
            self.content_label.pack(pady=10)

            scrollable_frame = ctk.CTkScrollableFrame(self.content_area, width=960, height=580, corner_radius=0)
            scrollable_frame.pack(padx=10, pady=10, fill="both", expand=True)

            columns = 5  # 5 cards per row

            # Load products from Excel
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Products"]
                products = list(ws.iter_rows(min_row=2, values_only=True))  # skip header
            except Exception as e:
                print(f"Error loading Excel file: {e}")
                products = []

            for idx, (name, price, stock, image_path) in enumerate(products):
                row = idx // columns
                col = idx % columns

                card = ctk.CTkFrame(scrollable_frame, width=300, height=320, fg_color="#1e1e1e", corner_radius=15)
                card.grid(row=row, column=col, padx=20, pady=20)

                # Load product image
                try:
                    if image_path and os.path.exists(image_path):
                        img = Image.open(image_path)
                    elif os.path.exists("default.png"):
                        img = Image.open("default.png")
                    else:
                        raise FileNotFoundError("No image found.")

                    img = img.resize((100, 100), Image.LANCZOS)
                    photo = ImageTk.PhotoImage(img)

                except Exception as e:
                    print(f"Image load error: {e}")
                    img = Image.new("RGB", (220, 120), color="gray")
                    photo = ImageTk.PhotoImage(img)

                image_label = ctk.CTkLabel(card, text="", image=photo)
                image_label.image = photo
                image_label.pack(pady=(10, 5))

                name_label = ctk.CTkLabel(card, text=name, font=("Arial", 14, "bold"), wraplength=240)
                name_label.pack(pady=(0, 2))

                price_label = ctk.CTkLabel(card, text=f"₱{price:.2f}", font=("Arial", 13), text_color="#00e676")
                price_label.pack()

                stock_label = ctk.CTkLabel(card, text=f"Stock: {stock}", font=("Arial", 12))
                stock_label.pack(pady=(0, 5))

                product_obj = Product(name, price, stock, image_path)

                add_btn = ctk.CTkButton(
                    card,
                    text="Add to Cart",
                    width=140,
                    command=lambda p=product_obj: self.add_to_cart(p))
                add_btn.pack()

        def show_sale():
            set_active(sale_btn)
            clear_content()

            self.content_label = ctk.CTkLabel(self.content_area, text="💸 Sales Page", font=ctk.CTkFont(size=18))
            self.content_label.pack(pady=10)

            columns = ("Date", "Product", "Quantity", "Price")
            tree = ttk.Treeview(self.content_area, columns=columns, show="headings")
            for col in columns:
                tree.heading(col, text=col)
                tree.column(col, anchor="center", width=150)
            tree.pack(fill="both", expand=True)

            for sale in sales:
                tree.insert("", tk.END, values=sale)

        def show_about():
            set_active(about_btn)
            clear_content()

            # Title
            self.content_label = ctk.CTkLabel(self.content_area, text="📘 About Us", text_color='#FFFFFF', font=ctk.CTkFont(size=26, weight="bold"))
            self.content_label.pack(pady=(10, 5))

            # Intro
            intro = (
                "The One-stop Shop is a school project built by passionate student developers. "
                "Our system provides a modern shopping experience with an easy to use and configure UI/UX. "
                "Enjoy!"
            )
            intro_label = ctk.CTkLabel(self.content_area, text=intro,text_color='#FFFFFF', font=("Arial", 14), wraplength=900, justify="center")
            intro_label.pack(pady=10)

            # Frame for team members
            team_frame = ctk.CTkFrame(self.content_area, corner_radius=10)
            team_frame.pack(padx=40, pady=20, fill="both", expand=True)

            # Scrollable canvas
            canvas = tk.Canvas(team_frame, bg="#1a1a1a", bd=0, highlightthickness=0)
            scrollbar = ctk.CTkScrollbar(team_frame, orientation="vertical", command=canvas.yview)
            scrollable_frame = ctk.CTkFrame(canvas)

            scrollable_frame.bind("<Configure>", lambda e: canvas.configure(scrollregion=canvas.bbox("all")))
            canvas.create_window((0, 0), window=scrollable_frame, anchor="nw")
            canvas.configure(yscrollcommand=scrollbar.set)

            canvas.pack(side="left", fill="both", expand=True)
            scrollbar.pack(side="right", fill="y")

            # Team members and roles
            team = [
                ("James Miro R. Orte", "👨‍💼 Project Leader"),
                ("Markie Louie Catungal", "💻 Main Developer"),
                ("Angela Lipaopao", "🎨 UI/UX Designer"),
                ("Angeline Magada", "🖼️ Frontend Developer"),
                ("Ashly Joy Tubiera", "🧪 Quality Assurance"),
                ("Christian Talavera", "🧠 Backend Developer"),
                ("Cris Laurence Pelaez", "📦 Inventory & DB Manager"),
                ("John Derick Gaviola", "🧾 Documentation Lead"),
                ("Elieza Tiay", "🔍 Bug Testing & Feedback"),
                ("John Irvin Suarez", "📱 Frontend Developer"),
                ("Jelord Willan Bajade", "🌐 Backend Developer"),
                ("Justine Rodriguez", "📑 Feature Planner"),
                ("Khian Mharc Pineda", "🛒 Cart & Checkout Logic"),
                ("Kimberly Portez", "📝 Report & Presentation"),
                ("Kyzel Roperez", "📸 Media & Assets"),
                ("Lord Raphael Lorena", "📤 Deployment & Packaging"),
            ]

            for name, role in team:
                row = ctk.CTkFrame(scrollable_frame, fg_color="transparent")
                row.pack(fill="x", padx=20, pady=5)

                name_label = ctk.CTkLabel(row, text=name, width=300, anchor="w",text_color='#00dc00', font=("Arial", 14, "bold"))
                name_label.pack(side="left")

                role_label = ctk.CTkLabel(row, text=role, anchor="w",text_color='#00dc00' ,font=("Arial", 14))
                role_label.pack(side="right")

            # Footer
            footer = ctk.CTkLabel(
                self.content_area,
                text="One-stop Shop • Version 1.0.0 • © 2025 | For Academic Use Only",
                font=("Arial", 12),
                text_color="gray"
            )
            footer.pack(pady=15)

        def show_cart():
            set_active(cart_btn)
            clear_content()

            self.content_label = ctk.CTkLabel(self.content_area, text="🛒 Your Cart", font=ctk.CTkFont(size=18))
            self.content_label.pack(pady=10)

            scrollable_frame_cart = ctk.CTkScrollableFrame(self.content_area, width=960, height=480, corner_radius=0)
            scrollable_frame_cart.pack(padx=10, pady=10, fill="both", expand=True)

            if not self.current_user.cart:
                empty_label = ctk.CTkLabel(scrollable_frame_cart, text="Your cart is empty.", text_color="gray")
                empty_label.pack(pady=20)
                return

            total = 0

            for idx, product in enumerate(self.current_user.cart):
                frame = ctk.CTkFrame(scrollable_frame_cart, fg_color="#2b2b2b", corner_radius=10)
                frame.pack(fill="x", padx=10, pady=5)

                name_label = ctk.CTkLabel(frame, text=product.name, font=("Arial", 14))
                name_label.pack(anchor="w", padx=5, pady=5)

                price_label = ctk.CTkLabel(frame, text=f"₱{product.price:.2f}", font=("Arial", 12))
                price_label.pack(anchor="w", padx=5)

                remove_btn = ctk.CTkButton(frame, text="🗑️ Remove", fg_color="red", width=80,
                                        command=lambda p=product: remove_from_cart(p))
                remove_btn.pack(anchor="e", padx=5, pady=5)

                total += product.price

            total_label = ctk.CTkLabel(self.content_area, text=f"Total: ₱{total:.2f}", font=("Arial", 16, "bold"))
            total_label.place(x=600,y=10)

            checkout_btn = ctk.CTkButton(self.content_area, text="Checkout", command=self.checkout)
            checkout_btn.place(x=800,y=10)


        def remove_from_cart(product):
            self.current_user.cart.remove(product)
            self.show_cart()  # refresh

        product_btn = ctk.CTkButton(toolbar, text="Products", command=show_products)
        product_btn.pack(side="left", padx=5, pady=5)
        nav_buttons.append(product_btn)

        sale_btn = ctk.CTkButton(toolbar, text="Sale", command=show_sale)
        sale_btn.pack(side="left", padx=5, pady=5)
        nav_buttons.append(sale_btn)

        about_btn = ctk.CTkButton(toolbar, text="About Us", command=show_about)
        about_btn.pack(side="left", padx=5, pady=5)
        nav_buttons.append(about_btn)

        cart_btn = ctk.CTkButton(toolbar, text="Cart 🛒", command=show_cart)
        cart_btn.pack(side="left", padx=5, pady=5)
        nav_buttons.append(cart_btn)

        def exit():
            if messagebox.askokcancel("Exit", "Are you sure you want to exit?") == False:
                pass
            else:
                self.user_login_screen()

        exit_btn = ctk.CTkButton(toolbar, text="Exit", command=exit)
        exit_btn.pack(side="right", padx=5, pady=5)

    def add_to_cart(self, product):
        if product.stock > 0:
            self.current_user.cart.append(product)
            product.stock -= 1

            # Update stock in Excel
            try:
                wb = load_workbook(EXCEL_FILE)
                ws = wb["Products"]
                for row in ws.iter_rows(min_row=2):
                    if row[0].value == product.name:
                        row[2].value = product.stock  # Update stock column (C)
                        break
                wb.save(EXCEL_FILE)
            except Exception as e:
                messagebox.showerror("Excel Error", f"Failed to update stock: {e}")
                return

            self.show_products()  # 🔁 Immediately refresh products page
        else:
            messagebox.showerror("Out of Stock", "This product is no longer available.")

    def checkout(self):
        if not self.current_user:
            messagebox.showerror("Not Logged In", "Please log in to proceed with checkout.")
            return

        if not self.current_user.cart:
            messagebox.showwarning("Cart Empty", "Your cart is empty.")
            return

        # Add cart items to purchase history
        self.current_user.purchase_history.extend(self.current_user.cart)

        # Save receipt (simple .txt version)
        try:
            receipt_text = f'Thank you for using the One-stop Shop For All! Please come again\n'
            receipt_text += f"Receipt for {self.current_user.username}\n"
            receipt_text += "--------------------------------------\n"
            total = 0
            for item in self.current_user.cart:
                line = f"{item.name} - ₱{item.price:.2f}\n"
                receipt_text += line
                total += item.price

                sales.append([self.current_user.username, item.name, item.price, datetime.now().strftime("%Y-%m-%d %H:%M:%S")])

            receipt_text += "--------------------------------------\n"
            receipt_text += f"Total: ₱{total:.2f}\n"

            # Save to file
            with open(f"receipt_{self.current_user.username}.txt", "w", encoding="utf-8") as f:
                f.write(receipt_text)

        except Exception as e:
            messagebox.showerror("Receipt Error", f"Failed to generate receipt: {e}")
            return
        
        # Clear the cart
        self.current_user.cart.clear()
        

        messagebox.showinfo("Checkout Successful", f"Checkout complete! Receipt saved as receipt_{self.current_user.username}.txt")

        # Optional: refresh cart UI if needed
        return self.show_cart()

# Run App
if __name__ == "__main__":
    app = ECommerceApp()
    app.mainloop()