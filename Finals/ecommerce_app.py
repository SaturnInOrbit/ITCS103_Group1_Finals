import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
from openpyxl import load_workbook
import os

EXCEL_FILE = "inventory.xlsx"

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Name", "Price", "Stock", "Image"])
        wb.save(EXCEL_FILE)

ensure_excel_file()

# Setup
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")  # We'll manually adjust purple later

# Constants
ADMIN_USER = "admin"
ADMIN_PASS = "password"

# Global storage
inventory = []
sales = []
purchase_history = []
user_accounts = {}  # Dictionary to store user credentials and data

# Product structure
class Product:
    def __init__(self, name, price, stock, image_path):
        self.name = name
        self.price = price
        self.stock = stock
        self.image_path = image_path

class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.cart = []
        self.purchase_history = []

# Main App
class ECommerceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("E-Commerce Inventory System")
        self.geometry("1000x700")
        self.resizable(False, False)
        self.configure(bg="#1a1a1a")

        self.error_label = None
        self.selected_product = None
        self.current_user = None

        self.home_screen()

    def clear_widgets(self):
        for widget in self.winfo_children():
            widget.destroy()

    def home_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="Welcome to the E-Commerce System", font=("Arial", 30), text_color="#b84dff")
        title.pack(pady=30)

        admin_btn = ctk.CTkButton(frame, text="Admin Login", command=self.admin_login_screen, fg_color="#7a00cc", hover_color="#660099")
        admin_btn.pack(pady=20, ipadx=20)

        user_btn = ctk.CTkButton(frame, text="Shop Now", command=self.user_login_screen, fg_color="#00cc66", hover_color="#00994d")
        user_btn.pack(pady=20, ipadx=30)
    
    def admin_login_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="Admin Login", font=("Arial", 26), text_color="#b84dff")
        title.pack(pady=20)

        user_label = ctk.CTkLabel(frame, text="Username:")
        user_label.pack(pady=(10, 0))
        self.admin_username = ctk.CTkEntry(frame)
        self.admin_username.pack(pady=(0, 10))

        pass_label = ctk.CTkLabel(frame, text="Password:")
        pass_label.pack(pady=(10, 0))
        self.admin_password = ctk.CTkEntry(frame, show="*")
        self.admin_password.pack(pady=(0, 10))

        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_admin)
        login_btn.pack(pady=10)

        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen)
        back_btn.pack(pady=10)

    def verify_admin(self):
        username = self.admin_username.get()
        password = self.admin_password.get()

        if username == ADMIN_USER and password == ADMIN_PASS:
            self.admin_dashboard()
        else:
            messagebox.showerror("Login Failed", "Invalid admin credentials")
    
# --- ADMIN DASHBOARD ---
    def admin_dashboard(self):
        self.clear_widgets()

        title = ctk.CTkLabel(self, text="Admin Dashboard", font=("Arial", 28), text_color="#ff9933")
        title.pack(pady=20)

        add_product_btn = ctk.CTkButton(self, text="Add Product", command=self.add_product_popup)
        add_product_btn.pack(pady=10)

        download_sales_btn = ctk.CTkButton(self, text="Download Sales Report", command=self.download_sales_report)
        download_sales_btn.pack(pady=10)

        logout_btn = ctk.CTkButton(self, text="Logout", command=self.home_screen)
        logout_btn.pack(pady=10)

        self.admin_frame = ctk.CTkFrame(self, fg_color="#262626", height=400, width=900)
        self.admin_frame.pack(pady=20)

        self.refresh_inventory()

    def refresh_inventory(self):
        
        # Clear previous widgets
        for widget in self.admin_frame.winfo_children():
            widget.destroy()

        # Clear current inventory list
        inventory.clear()

        # Load from Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, stock, image_path = row
            if os.path.exists(image_path):  # Ensure image still exists
                product = Product(name, float(price), int(stock), image_path)
                inventory.append(product)

        if not inventory:
            label = ctk.CTkLabel(self.admin_frame, text="No Products Yet.", text_color="gray")
            label.pack()
            return

        # Display products
        for idx, product in enumerate(inventory):
            frame = ctk.CTkFrame(self.admin_frame, fg_color="#333333", corner_radius=10)
            frame.grid(row=idx//2, column=idx%2, padx=10, pady=10, sticky="nsew")

            try:
                img = Image.open(product.image_path)
                img = img.resize((100, 100))
                img = ImageTk.PhotoImage(img)
                img_label = ctk.CTkLabel(frame, image=img, text="")
                img_label.image = img  # keep reference
                img_label.pack()
            except Exception as e:
                error_label = ctk.CTkLabel(frame, text="Image error", text_color="red")
                error_label.pack()

            info = f"{product.name}\n₱{product.price:.2f}\nStock: {product.stock}"
            label = ctk.CTkLabel(frame, text=info)
            label.pack()

            edit_btn = ctk.CTkButton(frame, text="Edit Stock", command=lambda p=product: self.edit_stock_popup(p))
            edit_btn.pack(pady=2)   

            delete_btn = ctk.CTkButton(frame, text="Delete", fg_color="red", command=lambda p=product: self.delete_product(p))
            delete_btn.pack(pady=2)


    def add_product_popup(self):
        popup = tk.Toplevel(self)
        popup.title("Add Product")
        popup.geometry("300x400")
        popup.configure(bg="#1a1a1a")

        name_label = ctk.CTkLabel(popup, text="Product Name:")
        name_label.pack()
        name_entry = ctk.CTkEntry(popup)
        name_entry.pack()

        price_label = ctk.CTkLabel(popup, text="Price:")
        price_label.pack()
        price_entry = ctk.CTkEntry(popup)
        price_entry.pack()

        stock_label = ctk.CTkLabel(popup, text="Stock:")
        stock_label.pack()
        stock_entry = ctk.CTkEntry(popup)
        stock_entry.pack()

        image_btn = ctk.CTkButton(popup, text="Select Image", command=lambda: self.select_image(popup))
        image_btn.pack(pady=10)

        add_btn = ctk.CTkButton(popup, text="Add", command=lambda: self.save_product(name_entry.get(), price_entry.get(), stock_entry.get(), self.selected_image, popup))
        add_btn.pack(pady=10)

    def select_image(self, popup):
        path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg")])
        if path:
            self.selected_image = path
            messagebox.showinfo("Image Selected", "Image successfully selected.")

    def save_product(self, name, price, stock, image, popup):
        if not (name and price and stock and image):
            messagebox.showerror("Error", "All fields are required.")
            return
        try:
            price = float(price)
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Price must be a number. Stock must be an integer.")
            return

        new_product = Product(name, price, stock, image)
        inventory.append(new_product)

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]
        ws.append([name, price, stock, image])
        wb.save(EXCEL_FILE)

        popup.destroy()
        self.refresh_inventory()

    def edit_stock_popup(self, product):
        popup = tk.Toplevel(self)
        popup.title("Edit Stock")
        popup.geometry("250x200")
        popup.configure(bg="#1a1a1a")

        label = ctk.CTkLabel(popup, text=f"Editing stock for {product.name}")
        label.pack(pady=10)

        stock_entry = ctk.CTkEntry(popup)
        stock_entry.pack()

        save_btn = ctk.CTkButton(popup, text="Save", command=lambda: self.save_stock(product, stock_entry.get(), popup))
        save_btn.pack(pady=10)

    def save_stock(self, product, stock, popup):
        try:
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Stock must be an integer.")
            return

        # Update in-memory
        product.stock = stock

        # Update in Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product.name and row[1].value == product.price and row[3].value == product.image:
                row[2].value = stock
                break
        wb.save(EXCEL_FILE)

        popup.destroy()
        self.refresh_inventory()


    def delete_product(self, product):
        if messagebox.askyesno("Delete", f"Are you sure you want to delete {product.name}?"):
            inventory.remove(product)

            # Delete from Excel
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Products"]
            for row in list(ws.iter_rows(min_row=2)):
                if row[0].value == product.name and row[1].value == product.price and row[3].value == product.image:
                    ws.delete_rows(row[0].row)
                    break
            wb.save(EXCEL_FILE)

            self.refresh_inventory()


    def download_sales_report(self):
        if not sales:
            messagebox.showerror("Error", "No sales to report.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        headers = ["Username", "Product", "Price", "Date"]
        ws.append(headers)

        for sale in sales:
            ws.append(sale)

        wb.save(filepath)
        messagebox.showinfo("Success", f"Sales report saved to {filepath}.")

    def user_login_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="User Login", font=("Arial", 26), text_color="#00cc66")
        title.pack(pady=20)

        user_label = ctk.CTkLabel(frame, text="Username:")
        user_label.pack(pady=(10, 0))
        self.user_username = ctk.CTkEntry(frame)
        self.user_username.pack(pady=(0, 10))

        pass_label = ctk.CTkLabel(frame, text="Password:")
        pass_label.pack(pady=(10, 0))
        self.user_password = ctk.CTkEntry(frame, show="*")
        self.user_password.pack(pady=(0, 10))

        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_user)
        login_btn.pack(pady=10)

        register_btn = ctk.CTkButton(frame, text="Register", command=self.user_register_screen)
        register_btn.pack(pady=10)

        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen)
        back_btn.pack(pady=10)

    def verify_user(self):
        username = self.user_username.get()
        password = self.user_password.get()

        if username in user_accounts and user_accounts[username].password == password:
            self.current_user = user_accounts[username]
            self.user_shop_screen()
        else:
            messagebox.showerror("Login Failed", "Invalid username or password")

    def user_register_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="User Registration", font=("Arial", 26), text_color="#00cc66")
        title.pack(pady=20)

        user_label = ctk.CTkLabel(frame, text="New Username:")
        user_label.pack(pady=(10, 0))
        self.reg_username = ctk.CTkEntry(frame)
        self.reg_username.pack(pady=(0, 10))

        pass_label = ctk.CTkLabel(frame, text="New Password:")
        pass_label.pack(pady=(10, 0))
        self.reg_password = ctk.CTkEntry(frame, show="*")
        self.reg_password.pack(pady=(0, 10))

        register_btn = ctk.CTkButton(frame, text="Register", command=self.register_user)
        register_btn.pack(pady=20)

        back_btn = ctk.CTkButton(frame, text="Back", command=self.user_login_screen)
        back_btn.pack(pady=10)

    def register_user(self):
        username = self.reg_username.get()
        password = self.reg_password.get()

        if username in user_accounts:
            messagebox.showerror("Error", "Username already exists")
            return

        user_accounts[username] = User(username, password)
        messagebox.showinfo("Success", "Registration successful! Please login.")
        self.user_login_screen()

    def user_shop_screen(self):
        self.clear_widgets()

        sidebar = ctk.CTkFrame(self, width=200, fg_color="#333333")
        sidebar.pack(side="left", fill="y")

        content = ctk.CTkFrame(self, fg_color="#1a1a1a")
        content.pack(expand=True, fill="both")

        view_cart_btn = ctk.CTkButton(sidebar, text="View Cart", command=self.view_cart)
        view_cart_btn.pack(pady=20)

        history_btn = ctk.CTkButton(sidebar, text="Purchase History", command=self.view_purchase_history)
        history_btn.pack(pady=20)

        logout_btn = ctk.CTkButton(sidebar, text="Logout", command=self.home_screen)
        logout_btn.pack(pady=20)

        self.user_content = content
        self.show_products_for_user()

    def show_products_for_user(self):
        for widget in self.user_content.winfo_children():
            widget.destroy()

        title = ctk.CTkLabel(self.user_content, text="Shop Products", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        for product in inventory:
            block = ctk.CTkFrame(self.user_content, fg_color="#2b2b2b")
            block.pack(padx=10, pady=10, fill="x")

            name_label = ctk.CTkLabel(block, text=f"{product.name}", font=("Arial", 16))
            name_label.pack(anchor="w", padx=10)

            price_label = ctk.CTkLabel(block, text=f"₱{product.price:.2f}", font=("Arial", 14))
            price_label.pack(anchor="w", padx=10)

            stock_label = ctk.CTkLabel(block, text=f"Stock: {product.stock}", font=("Arial", 14))
            stock_label.pack(anchor="w", padx=10)

            add_btn = ctk.CTkButton(block, text="Add to Cart", command=lambda p=product: self.add_to_cart(p))
            add_btn.pack(anchor="e", padx=10)

    def add_to_cart(self, product):
        if product.stock > 0:
            self.current_user.cart.append(product)
            product.stock -= 1
            self.show_products_for_user()
        else:
            messagebox.showerror("Out of Stock", "This product is no longer available.")

    def view_cart(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#1a1a1a")
        frame.pack(expand=True, fill="both")

        title = ctk.CTkLabel(frame, text="Your Cart", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        total = 0
        for item in self.current_user.cart:
            label = ctk.CTkLabel(frame, text=f"{item.name} - ₱{item.price:.2f}", font=("Arial", 14))
            label.pack(anchor="w", padx=20)
            total += item.price

        total_label = ctk.CTkLabel(frame, text=f"Total: ₱{total:.2f}", font=("Arial", 16))
        total_label.pack(pady=10)

        checkout_btn = ctk.CTkButton(frame, text="Checkout", command=self.checkout)
        checkout_btn.pack(pady=10)

        back_btn = ctk.CTkButton(frame, text="Back to Shop", command=self.user_shop_screen)
        back_btn.pack(pady=10)

    def checkout(self):
        if not self.current_user.cart:
            messagebox.showerror("Empty Cart", "Your cart is empty!")
            return

        total_price = sum(item.price for item in self.current_user.cart)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for item in self.current_user.cart:
            sales.append([now, item.name, 1, item.price])
            self.current_user.purchase_history.append([now, item.name, 1, item.price])

        self.generate_receipt(self.current_user.cart, total_price)
        self.current_user.cart = []
        messagebox.showinfo("Success", "Purchase complete! Receipt downloaded.")
        self.user_shop_screen()

    def generate_receipt(self, items, total):
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipt"

        ws.append(["Product Name", "Price"])
        for item in items:
            ws.append([item.name, item.price])

        ws.append(["Total", total])

        filename = f"receipt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

    def view_purchase_history(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#1a1a1a")
        frame.pack(expand=True, fill="both")

        title = ctk.CTkLabel(frame, text="Purchase History", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        for record in self.current_user.purchase_history:
            label = ctk.CTkLabel(frame, text=f"{record[0]} - {record[1]} - ₱{record[3]:.2f}", font=("Arial", 14))
            label.pack(anchor="w", padx=20)

        back_btn = ctk.CTkButton(frame, text="Back to Shop", command=self.user_shop_screen)
        back_btn.pack(pady=10)

# Run App
if __name__ == "__main__":
    app = ECommerceApp()
    app.mainloop()
