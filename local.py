
import customtkinter as ctk
import tkinter as tk
from tkinter import filedialog, messagebox, ttk
from PIL import Image, ImageTk
import openpyxl
from openpyxl import Workbook
from openpyxl.styles import Alignment
from datetime import datetime
import openpyxl
from openpyxl import load_workbook
import os
from openpyxl.styles import Font

EXCEL_FILE = "inventory.xlsx"

def ensure_excel_file():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Products"
        ws.append(["Name", "Price", "Stock", "Image"])
        wb.save(EXCEL_FILE)

ensure_excel_file()

# Setup
ctk.set_appearance_mode("dark")
ctk.set_default_color_theme("green")  # We'll manually adjust purple later

# Constants
ADMIN_USER = "admin"
ADMIN_PASS = "password"
FILE_NAME = "users.xlsx"
# Global storage
inventory = []
sales = []
purchase_history = []
user_accounts = {}  # Dictionary to store user credentials and data

nav_buttons = []



# Product structure
class Product:
    def __init__(self, name, price, stock, image_path):
        self.name = name
        self.price = price
        self.stock = stock
        self.image_path = image_path

class User:
    def __init__(self, username, password):
        self.username = username
        self.password = password
        self.cart = []
        self.purchase_history = []



def user_window(self):
    self.clear_widgets()

    title_label = ctk.CTkLabel(self, text="🛍️ E-Commerce ni miko tanginanyo lahat", font=ctk.CTkFont(size=22, weight="bold"))
    title_label.pack(pady=(15, 5))

    # ---- Toolbar Frame (Horizontal Buttons) ----
    toolbar = ctk.CTkFrame(self)
    toolbar.pack(fill="x", pady=5)

    # ---- Content Area ----
    self.content_area = ctk.CTkFrame(self)
    self.content_area.pack(fill="both", expand=True, padx=20, pady=20)

    self.content_label = ctk.CTkLabel(self.content_area, text="Welcome to the E-Commerce Dashboard", font=ctk.CTkFont(size=18))
    self.content_label.pack(pady=20)

    def clear_content():
        for widget in self.content_area.winfo_children():
            widget.destroy()

    def set_active(button):
        for btn in nav_buttons:
            
            btn.configure(fg_color="transparent")
        button.configure(fg_color="#1a1a1a")

    def show_products():
        set_active(product_btn)
        clear_content()

        self.content_label = ctk.CTkLabel(self.content_area, text="📦 Products Page", font=ctk.CTkFont(size=18))
        self.content_label.pack(pady=10)

        columns = ("Product Name", "Price", "Stock")
        tree = ttk.Treeview(self.content_area, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=150)
        tree.pack(fill="both", expand=True)

        # Load data from inventory list
        for product in inventory:
            tree.insert("", tk.END, values=(product.name, f"₱{product.price:.2f}", product.stock))

    def show_sale():
        set_active(sale_btn)
        clear_content()

        self.content_label = ctk.CTkLabel(self.content_area, text="💸 Sales Page", font=ctk.CTkFont(size=18))
        self.content_label.pack(pady=10)

        columns = ("Date", "Product", "Quantity", "Price")
        tree = ttk.Treeview(self.content_area, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="center", width=150)
        tree.pack(fill="both", expand=True)

        for sale in sales:
            tree.insert("", tk.END, values=sale)

    def show_about():
        set_active(about_btn)
        clear_content()

        self.content_label = ctk.CTkLabel(self.content_area, text="📘 About Us Page", font=ctk.CTkFont(size=18))
        self.content_label.pack(pady=10)

        about_data = [
            ("Company", "E-Commerce ni Miko"),
            ("Developer", "superjamies"),
            ("Developer", "supermiko"),
            ("Version", "1.0.0"),
            ("Year", "2025")
        ]

        columns = ("Field", "Information")
        tree = ttk.Treeview(self.content_area, columns=columns, show="headings")
        for col in columns:
            tree.heading(col, text=col)
            tree.column(col, anchor="w", width=200)
        tree.pack(fill="both", expand=True)

        for item in about_data:
            tree.insert("", tk.END, values=item)


    def show_cart():
        set_active(cart_btn)
        clear_content()

        self.content_label = ctk.CTkLabel(self.content_area, text="🛒 Your Cart", font=ctk.CTkFont(size=18))
        self.content_label.pack(pady=10)

        if not self.current_user.cart:
            empty_label = ctk.CTkLabel(self.content_area, text="Your cart is empty.", text_color="gray")
            empty_label.pack(pady=20)
            return

        total = 0

        for idx, product in enumerate(self.current_user.cart):
            frame = ctk.CTkFrame(self.content_area, fg_color="#2b2b2b", corner_radius=10)
            frame.pack(fill="x", padx=20, pady=10)

            name_label = ctk.CTkLabel(frame, text=product.name, font=("Arial", 14))
            name_label.pack(anchor="w", padx=10, pady=5)

            price_label = ctk.CTkLabel(frame, text=f"₱{product.price:.2f}", font=("Arial", 12))
            price_label.pack(anchor="w", padx=10)

            remove_btn = ctk.CTkButton(frame, text="🗑️ Remove", fg_color="red", width=80,
                                    command=lambda p=product: remove_from_cart(p))
            remove_btn.pack(anchor="e", padx=10, pady=5)

            total += product.price

        total_label = ctk.CTkLabel(self.content_area, text=f"Total: ₱{total:.2f}", font=("Arial", 16, "bold"))
        total_label.pack(pady=10)

        checkout_btn = ctk.CTkButton(self.content_area, text="Checkout", command=self.checkout)
        checkout_btn.pack(pady=10)

    def remove_from_cart(product):
        self.current_user.cart.remove(product)
        self.show_cart()  # refresh

    product_btn = ctk.CTkButton(toolbar, text="Products", command=show_products)
    product_btn.pack(side="left", padx=5, pady=5)
    nav_buttons.append(product_btn)

    sale_btn = ctk.CTkButton(toolbar, text="Sale", command=show_sale)
    sale_btn.pack(side="left", padx=5, pady=5)
    nav_buttons.append(sale_btn)

    about_btn = ctk.CTkButton(toolbar, text="About Us", command=show_about)
    about_btn.pack(side="left", padx=5, pady=5)
    nav_buttons.append(about_btn)

    cart_btn = ctk.CTkButton(toolbar, text="Cart 🛒", command=show_cart)
    cart_btn.pack(side="left", padx=5, pady=5)
    nav_buttons.append(cart_btn)

    def exit():
        if messagebox.askyesno("Exit", "Are you sure you want to exit?"):
            self.destroy()

    exit_btn = ctk.CTkButton(toolbar, text="Exit", command=exit)
    exit_btn.pack(side="right", padx=5, pady=5)
        





def validate_login(username, password):
        wb = openpyxl.load_workbook(FILE_NAME)
        ws = wb.active

        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] == username and row[1] == password:
                return True
        return False

# Main App
class ECommerceApp(ctk.CTk):
    def __init__(self):
        super().__init__()
        self.title("E-Commerce Inventory System")
        self.geometry("1000x700")
        self.resizable(False, False)
        self.configure(bg="#1a1a1a")

        self.error_label = None
        self.selected_product = None
        self.current_user = None

        self.home_screen()

    def clear_widgets(self):
        for widget in self.winfo_children():
            widget.destroy()

    def home_screen(self):
        self.clear_widgets()

        # Main Frame
        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        # Title Label with updated font and color
        title = ctk.CTkLabel(frame, text="Welcome to the E-Commerce System", font=("Arial", 34, "bold"), text_color="#b84dff")
        title.pack(pady=30)

        # Admin Login Button with modern styling
        admin_btn = ctk.CTkButton(frame, text="👨‍💼 Admin Login", command=self.admin_login_screen, 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 16, "bold"))
        admin_btn.pack(pady=20, ipadx=30, ipady=10, fill="x", padx=40)

        # Shop Now Button with modern styling and added icon
        user_btn = ctk.CTkButton(frame, text="🛒 Shop Now", command=self.user_login_screen, 
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"))
        user_btn.pack(pady=20, ipadx=30, ipady=10, fill="x", padx=40)

    
    def admin_login_screen(self):
        self.clear_widgets()

        # Main Frame
        frame = ctk.CTkFrame(self, fg_color="#2a2a2a")
        frame.pack(expand=True, padx=40, pady=40)

        # Title with updated font size and color
        title = ctk.CTkLabel(frame, text="👨‍💼 Admin Login", font=("Arial", 26, "bold"), text_color="#b84dff")
        title.pack(pady=20,padx = 20)

        # Username Label and Entry
        user_label = ctk.CTkLabel(frame, text="Username:", font=("Arial", 14))
        user_label.pack(pady=(10, 0))
        self.admin_username = ctk.CTkEntry(frame, placeholder_text="Enter your username", font=("Arial", 14))
        self.admin_username.pack(pady=(0, 10),padx =10, fill="x")

        # Password Label and Entry
        pass_label = ctk.CTkLabel(frame, text="Password:", font=("Arial", 14))
        pass_label.pack(pady=(10, 0))
        self.admin_password = ctk.CTkEntry(frame, show="*", placeholder_text="Enter your password", font=("Arial", 14))
        self.admin_password.pack(pady=(0, 20),padx =10, fill="x")

        # Login Button with modern design
        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_admin, 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 16, "bold"), border_width=2, corner_radius=10)
        login_btn.pack(pady=(0, 10),padx= 10,  fill="x")

        # Back Button with updated design
        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen, 
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"), border_width=2, corner_radius=10)
        back_btn.pack(pady=10,padx =10, fill="x")


    def verify_admin(self):
        username = self.admin_username.get()
        password = self.admin_password.get()

        if username == ADMIN_USER and password == ADMIN_PASS:
            self.admin_dashboard()
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
        else:
            messagebox.showerror("Login Failed", "Invalid admin credentials")
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
    
    def admin_logout(self):
        messagebox.showwarning("Log Out", "Are You Sure You're Logging out")
        self.home_screen()
    
# --- ADMIN DASHBOARD ---
    def admin_dashboard(self):
        self.clear_widgets()

        # Main layout: sidebar and content area
        main_frame = ctk.CTkFrame(self)
        main_frame.pack(fill="both", expand=True)

        # Sidebar
        sidebar = ctk.CTkFrame(main_frame, width=200, fg_color="#1a1a1a")
        sidebar.pack(side="left", fill="y")

        title = ctk.CTkLabel(sidebar, text="👨‍💼 Admin", font=("Arial", 24, "bold"), text_color="#ff9933")
        title.pack(pady=(20, 30))

        # Sidebar buttons
        add_product_btn = ctk.CTkButton(sidebar, text="➕ Add Product", command=self.add_product_popup)
        add_product_btn.pack(pady=10, fill="x", padx=10)

        download_sales_btn = ctk.CTkButton(sidebar, text="📥 Download Sales", command=self.download_sales_report)
        download_sales_btn.pack(pady=10, fill="x", padx=10)

        logout_btn = ctk.CTkButton(sidebar, text="🚪 Logout", command=self.admin_logout)
        logout_btn.pack(side="bottom", pady=20, fill="x", padx=10)

        # Content Area
        self.admin_frame = ctk.CTkFrame(main_frame, fg_color="#262626")
        self.admin_frame.pack(side="right", fill="both", expand=True, padx=10, pady=10)

        self.refresh_inventory()


    def refresh_inventory(self):
        from PIL import Image, ImageTk
        import os

        # Clear previous widgets
        for widget in self.admin_frame.winfo_children():
            widget.destroy()

        # Clear current inventory list
        inventory.clear()

        # Load from Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]

        for row in ws.iter_rows(min_row=2, values_only=True):
            name, price, stock, image_path = row
            if os.path.exists(image_path):  # Ensure image still exists
                product = Product(name, float(price), int(stock), image_path)
                inventory.append(product)

        if not inventory:
            label = ctk.CTkLabel(self.admin_frame, text="No Products Yet.", text_color="gray")
            label.pack()
            return

        # Display products
        for idx, product in enumerate(inventory):
            frame = ctk.CTkFrame(self.admin_frame, fg_color="#333333", corner_radius=10)
            frame.grid(row=idx//2, column=idx%2, padx=10, pady=10, sticky="nsew")

            try:
                img = Image.open(product.image_path)
                img = img.resize((100, 100))
                img = ImageTk.PhotoImage(img)
                img_label = ctk.CTkLabel(frame, image=img, text="")
                img_label.image = img  # keep reference
                img_label.pack()
            except Exception as e:
                error_label = ctk.CTkLabel(frame, text="Image error", text_color="red")
                error_label.pack()

            info = f"{product.name}\n₱{product.price:.2f}\nStock: {product.stock}"
            label = ctk.CTkLabel(frame, text=info)
            label.pack()

            edit_btn = ctk.CTkButton(frame, text="Edit Stock", command=lambda p=product: self.edit_stock_popup(p))
            edit_btn.pack(pady=2)   

            delete_btn = ctk.CTkButton(frame, text="Delete", fg_color="red", command=lambda p=product: self.delete_product(p))
            delete_btn.pack(pady=2)


    def add_product_popup(self):
        # Create the popup window
        popup = tk.Toplevel(self)
        popup.title("Add Product")
        popup.geometry("350x450")
        popup.configure(bg="#2a2a2a")  # Darker background for better contrast

        # Product Name Label and Entry
        name_label = ctk.CTkLabel(popup, text="Product Name:", font=("Arial", 14), text_color="#b84dff")
        name_label.pack(pady=(20, 5))
        name_entry = ctk.CTkEntry(popup, placeholder_text="Enter product name", font=("Arial", 14))
        name_entry.pack(pady=(0, 10),padx=10, fill="x")

        # Price Label and Entry
        price_label = ctk.CTkLabel(popup, text="Price:", font=("Arial", 14), text_color="#b84dff")
        price_label.pack(pady=(10, 5))
        price_entry = ctk.CTkEntry(popup, placeholder_text="Enter product price", font=("Arial", 14))
        price_entry.pack(pady=(0, 10),padx=10, fill="x")

        # Stock Label and Entry
        stock_label = ctk.CTkLabel(popup, text="Stock:", font=("Arial", 14), text_color="#b84dff")
        stock_label.pack(pady=(10, 5))
        stock_entry = ctk.CTkEntry(popup, placeholder_text="Enter stock quantity", font=("Arial", 14))
        stock_entry.pack(pady=(0, 10),padx =10, fill="x") 

        # Image Button
        image_btn = ctk.CTkButton(popup, text="➕ Select Image", command=lambda: self.select_image(popup), 
                                fg_color="#7a00cc", hover_color="#660099", font=("Arial", 14))
        image_btn.pack(pady=(10, 20),padx = 10, fill="x")

        # Add Product Button
        add_btn = ctk.CTkButton(popup, text="Add Product", command=lambda: self.save_product(name_entry.get(), 
                                                                                            price_entry.get(), 
                                                                                            stock_entry.get(), 
                                                                                            self.selected_image, 
                                                                                            popup),
                                fg_color="#00cc66", hover_color="#00994d", font=("Arial", 16, "bold"))
        add_btn.pack(pady=(10, 20),padx =10, ipadx=30, ipady=10, fill="x")

        # Back Button (optional)
        back_btn = ctk.CTkButton(popup, text="Back", command=popup.destroy, 
                                fg_color="#ff3333", hover_color="#cc2900", font=("Arial", 14, "bold"))
        back_btn.pack(pady=(10, 10), ipadx=20, ipady=10, fill="x")


    def select_image(self, popup):
        path = filedialog.askopenfilename(filetypes=[("Image files", "*.png *.jpg *.jpeg")])
        if path:
            self.selected_image = path
            messagebox.showinfo("Image Selected", "Image successfully selected.")

    def save_product(self, name, price, stock, image, popup):
        if not (name and price and stock and image):
            messagebox.showerror("Error", "All fields are required.")
            return
        try:
            price = float(price)
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Price must be a number. Stock must be an integer.")
            return

        new_product = Product(name, price, stock, image)
        inventory.append(new_product)

        # Save to Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]
        ws.append([name, price, stock, image])
        wb.save(EXCEL_FILE)

        popup.destroy()
        self.refresh_inventory()

    def edit_stock_popup(self, product):
        popup = tk.Toplevel(self)
        popup.title("Edit Stock")
        popup.geometry("250x200")
        popup.configure(bg="#1a1a1a")

        label = ctk.CTkLabel(popup, text=f"Editing stock for {product.name}")
        label.pack(pady=10)

        stock_entry = ctk.CTkEntry(popup)
        stock_entry.pack()

        save_btn = ctk.CTkButton(popup, text="Save", command=lambda: self.save_stock(product, stock_entry.get(), popup))
        save_btn.pack(pady=10)

    def save_stock(self, product, stock, popup):
        try:
            stock = int(stock)
        except ValueError:
            messagebox.showerror("Error", "Stock must be an integer.")
            return

        # Update in-memory
        product.stock = stock

        # Update in Excel
        wb = load_workbook(EXCEL_FILE)
        ws = wb["Products"]
        for row in ws.iter_rows(min_row=2):
            if row[0].value == product.name and row[1].value == product.price and row[3].value == product.image:
                row[2].value = stock
                break
        wb.save(EXCEL_FILE)

        popup.destroy()
        self.refresh_inventory()


    def delete_product(self, product): ##not working
        if messagebox.askyesno("Delete", f"Are you sure you want to delete {product.name}?"):
            # Remove from in-memory list
            if product in inventory:
                inventory.remove(product)

            # Remove from Excel
            wb = load_workbook(EXCEL_FILE)
            ws = wb["Products"]

            for row in list(ws.iter_rows(min_row=2)):
                excel_name = str(row[0].value).strip()
                excel_price = float(row[1].value)
                excel_stock = int(row[2].value)
                excel_image = str(row[3].value).strip() if row[3].value else ""

                if (excel_name.lower() == product.name.lower() and
                    excel_price == product.price and
                    excel_stock == product.stock and
                    excel_image == product.image_path.strip()):
                    
                    ws.delete_rows(row[0].row)
                    break

            wb.save(EXCEL_FILE)

            # Optional: delete image file if it exists
            try:
                if os.path.exists(product.image_path):
                    os.remove(product.image_path)
            except Exception as e:
                print("Warning: Could not delete image file:", e)

            # Refresh product display
            self.refresh_inventory()




    def download_sales_report(self):
        if not sales:
            messagebox.showerror("Error", "No sales to report.")
            return

        filepath = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])

        if not filepath:
            return

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sales Report"

        headers = ["Username", "Product", "Price", "Date"]
        ws.append(headers)

        for sale in sales:
            ws.append(sale)

        wb.save(filepath)
        messagebox.showinfo("Success", f"Sales report saved to {filepath}.")

    def user_login_screen(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        
        title = ctk.CTkLabel(frame, text="User Login", font=("Arial", 26, "bold"), text_color="#00cc66")
        title.pack(pady=(10, 20))

        # Username fiel
        user_label = ctk.CTkLabel(frame, text="Username:", text_color="white", font=("Arial", 14))
        user_label.pack(anchor="w", padx=10)
        self.user_username = ctk.CTkEntry(frame, width=250, font=("Arial", 14), fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.user_username.pack(pady=(0, 10), padx=10)

        # Password fieldd
        pass_label = ctk.CTkLabel(frame, text="Password:", text_color="white", font=("Arial", 14))
        pass_label.pack(anchor="w", padx=10)
        self.user_password = ctk.CTkEntry(frame, width=250, font=("Arial", 14), show="*", fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.user_password.pack(pady=(0, 20), padx=10)

        # Login Button
        login_btn = ctk.CTkButton(frame, text="Login", command=self.verify_user, font=("Arial", 14), fg_color="#00cc66", hover_color="#00994d", width=250, corner_radius=10)
        login_btn.pack(pady=(10, 10), padx=10)

        # Register Button (if user doesn't have an account yet)
        register_btn = ctk.CTkButton(frame, text="Register", command=self.user_register_screen, font=("Arial", 14), fg_color="#b84dff", hover_color="#9933cc", width=250, corner_radius=10)
        register_btn.pack(pady=(5, 10), padx=10)

        # Back Button to tangina gulo nad
        back_btn = ctk.CTkButton(frame, text="Back", command=self.home_screen, font=("Arial", 14), fg_color="#ff6666", hover_color="#ff4d4d", width=250, corner_radius=10)
        back_btn.pack(pady=(5, 10), padx=10)

    
    
    def verify_user(self):
        username = self.user_username.get()
        password = self.user_password.get()

        if validate_login(username, password):
            messagebox.showinfo("Login Success", f"Welcome, {username}!")
            user_window(self)
            self.admin_username.delete(0, "end")
            self.admin_password.delete(0, "end")
        else:
            messagebox.showerror("Login Failed", "Invalid username or password.")
            self.user_username.delete(0, "end")
            self.user_password.delete(0, "end")




    def user_register_screen(self):
        self.clear_widgets()

        frame = ctk.CTkFrame(self, fg_color="#262626")
        frame.pack(expand=True)

        title = ctk.CTkLabel(frame, text="User Registration", font=("Arial", 26, "bold"), text_color="#00cc66")
        title.pack(pady=20)

        user_label = ctk.CTkLabel(frame, text="New Username:", font=("Arial", 14), text_color="white")
        user_label.pack(anchor="w", padx=10, pady=(10, 0))
        self.reg_username = ctk.CTkEntry(frame, width=250, font=("Arial", 14), fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.reg_username.pack(pady=(0, 10), padx=10)

        pass_label = ctk.CTkLabel(frame, text="New Password:", font=("Arial", 14), text_color="white")
        pass_label.pack(anchor="w", padx=10, pady=(10, 0))
        self.reg_password = ctk.CTkEntry(frame, width=250, font=("Arial", 14), show="*", fg_color="#333333", text_color="white", border_width=2, corner_radius=10)
        self.reg_password.pack(pady=(0, 10), padx=10)

        register_btn = ctk.CTkButton(frame, text="Register", command=self.save_register_user, font=("Arial", 14), fg_color="#00cc66", hover_color="#00994d", width=250, corner_radius=10)
        register_btn.pack(pady=(10, 10), padx=10)

        back_btn = ctk.CTkButton(frame, text="Back", command=self.user_login_screen, font=("Arial", 14), fg_color="#ff6666", hover_color="#ff4d4d", width=250, corner_radius=10)
        back_btn.pack(pady=(5, 10), padx=10)


    def register_user(self):
        if os.path.exists(FILE_NAME):
            wb = openpyxl.load_workbook(FILE_NAME)
            ws = wb.active
        else:
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "accounts"
            ws.append(["username", "password"])
            for col in range(1, 3):  # Fix range here
                ws.cell(row=1, column=col).font = Font(bold=True)
        return wb, ws

    def save_register_user(self):
        username = self.reg_username.get()
        password = self.reg_password.get()

        if not username or not password:
            messagebox.showerror("error","plesase fill up the entries")

        wb, ws = self.register_user()
        ws.append([username, password])
        wb.save(FILE_NAME)
          
        self.reg_username.delete(0, "end")
        self.reg_password.delete(0, "end")
        messagebox.showinfo(f"User {username}", "registered successfully.")

    
    def user_shop_screen(self):
        self.clear_widgets()

        sidebar = ctk.CTkFrame(self, width=200, fg_color="#333333")
        sidebar.pack(side="left", fill="y")

        content = ctk.CTkFrame(self, fg_color="#1a1a1a")
        content.pack(expand=True, fill="both")

        view_cart_btn = ctk.CTkButton(sidebar, text="View Cart", command=self.view_cart)
        view_cart_btn.pack(pady=20)

        history_btn = ctk.CTkButton(sidebar, text="Purchase History", command=self.view_purchase_history)
        history_btn.pack(pady=20)

        logout_btn = ctk.CTkButton(sidebar, text="Logout", command=self.home_screen)
        logout_btn.pack(pady=20)

        self.user_content = content
        self.show_products_for_user()

    def show_products_for_user(self):
        for widget in self.user_content.winfo_children():
            widget.destroy()

        title = ctk.CTkLabel(self.user_content, text="Shop Products", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        for product in inventory:
            block = ctk.CTkFrame(self.user_content, fg_color="#2b2b2b")
            block.pack(padx=10, pady=10, fill="x")

            name_label = ctk.CTkLabel(block, text=f"{product.name}", font=("Arial", 16))
            name_label.pack(anchor="w", padx=10)

            price_label = ctk.CTkLabel(block, text=f"₱{product.price:.2f}", font=("Arial", 14))
            price_label.pack(anchor="w", padx=10)

            stock_label = ctk.CTkLabel(block, text=f"Stock: {product.stock}", font=("Arial", 14))
            stock_label.pack(anchor="w", padx=10)

            add_btn = ctk.CTkButton(block, text="Add to Cart", command=lambda p=product: self.add_to_cart(p))
            add_btn.pack(anchor="e", padx=10)

    def add_to_cart(self, product):
        if product.stock > 0:
            self.current_user.cart.append(product)
            product.stock -= 1
            self.show_products_for_user()
        else:
            messagebox.showerror("Out of Stock", "This product is no longer available.")

    def view_cart(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#1a1a1a")
        frame.pack(expand=True, fill="both")

        title = ctk.CTkLabel(frame, text="Your Cart", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        total = 0
        for item in self.current_user.cart:
            label = ctk.CTkLabel(frame, text=f"{item.name} - ₱{item.price:.2f}", font=("Arial", 14))
            label.pack(anchor="w", padx=20)
            total += item.price

        total_label = ctk.CTkLabel(frame, text=f"Total: ₱{total:.2f}", font=("Arial", 16))
        total_label.pack(pady=10)

        checkout_btn = ctk.CTkButton(frame, text="Checkout", command=self.checkout)
        checkout_btn.pack(pady=10)

        back_btn = ctk.CTkButton(frame, text="Back to Shop", command=self.user_shop_screen)
        back_btn.pack(pady=10)

    def checkout(self):
        if not self.current_user.cart:
            messagebox.showerror("Empty Cart", "Your cart is empty!")
            return

        total_price = sum(item.price for item in self.current_user.cart)
        now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

        for item in self.current_user.cart:
            sales.append([now, item.name, 1, item.price])
            self.current_user.purchase_history.append([now, item.name, 1, item.price])

        self.generate_receipt(self.current_user.cart, total_price)
        self.current_user.cart = []
        messagebox.showinfo("Success", "Purchase complete! Receipt downloaded.")
        self.user_shop_screen()

    def generate_receipt(self, items, total):
        wb = Workbook()
        ws = wb.active
        ws.title = "Receipt"

        ws.append(["Product Name", "Price"])
        for item in items:
            ws.append([item.name, item.price])

        ws.append(["Total", total])

        filename = f"receipt_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        wb.save(filename)

    def view_purchase_history(self):
        self.clear_widgets()
        frame = ctk.CTkFrame(self, fg_color="#1a1a1a")
        frame.pack(expand=True, fill="both")

        title = ctk.CTkLabel(frame, text="Purchase History", font=("Arial", 24), text_color="#00cc66")
        title.pack(pady=20)

        for record in self.current_user.purchase_history:
            label = ctk.CTkLabel(frame, text=f"{record[0]} - {record[1]} - ₱{record[3]:.2f}", font=("Arial", 14))
            label.pack(anchor="w", padx=20)

        back_btn = ctk.CTkButton(frame, text="Back to Shop", command=self.user_shop_screen)
        back_btn.pack(pady=10)

# Run App
if __name__ == "__main__":
    app = ECommerceApp()
    app.mainloop()
